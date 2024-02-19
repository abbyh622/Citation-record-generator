from pathlib import Path
import os
import re
from bs4 import BeautifulSoup
import sqlite3
import pandas
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import tqdm

# default number of recent years to analyze
# change this to be input 
recent = 3
# need to know current year to know recent years
currentYear = datetime.now().year
recentYearEnd = currentYear - recent

# create database
Path('jngcitations.db').touch()
conn = sqlite3.connect('jngcitations.db')
cur = conn.cursor()

# create tables for articles and citations if not exists, delete previous data if exists
cur.execute('''CREATE TABLE IF NOT EXISTS articles (
            id TEXT PRIMARY KEY,
            volume INT,
            issue VARCHAR(3), 
            year INT,
            title TEXT,
            authors TEXT,
            totalCitations INT,
            recentCitations INT,
            citations TEXT
)''')
cur.execute('''CREATE TABLE IF NOT EXISTS citations (
            article TEXT,
            info TEXT,
            link TEXT,
            year INT,
            FOREIGN KEY (article) REFERENCES articles (id)
)''')
cur.execute('''DELETE FROM articles''')
cur.execute('''DELETE FROM citations''')

# get all file names in html temp folder
htmlVault = os.path.join(os.getcwd(), 'temp/htmls_temp')
fileList = os.listdir(htmlVault)

# patterns to get stuff we need
yearRegex = r'\b\d{4}\b'    
volRegex = r'\b\d+\b'           # just get number not 'Volume'
issRegex = r'\d+(?:-\d+)?'      # issue number is 1 number or 2 with dash

# progress bar for this?
# loop to extract the info from each file 
print("Extracting information...")
for filename in tqdm.tqdm(fileList):
    filePath = os.path.join(htmlVault, filename)
    with open(filePath, 'r', encoding = 'utf-8') as html: 
        # get title, authors, volume, issue, year, citations
        soup = BeautifulSoup(html, 'html.parser')
        volIss = soup.find(class_ = 'issue-heading')
        volIssue = [text for text in volIss.stripped_strings]
        title = soup.find('h1').text
        auths = soup.find_all(class_ = 'author')
        authors = [author.text for author in auths]
        cits = soup.find_all(class_ = 'citedByEntry')

        # get only the numbers        
        volYear = volIssue[0].split(',')
        if volYear[0] == 'Latest Articles':     # published online, will be in future volume
            volume = None
            issue = None
            year = None
        else:
            volume = re.search(volRegex, volYear[0]).group()
            issue = re.search(issRegex, volIssue[1]).group()
            year = re.search(yearRegex, volYear[1]).group()

        # list to contain all citations
        citations = []

        for cit in cits:
            # dict for details of each citation
            # want each citation to have the whole text in normal citation format, extract year
            # for queries, and extract link so it can be in a separate column for easier access
            citation = {}

            # strip white space and newline chars from text
            strings = list(cit.stripped_strings)
            text = ' '.join(strings).strip().replace('\n', '')
            # i just think this is better idk
            text = text.replace('Crossref', '(Crossref)')

            # put whole text into citation
            citation['info'] = text

            # extract year from text
            hasYear = re.search(yearRegex, text)
            if hasYear:
                cyear = hasYear.group()
                citation['year'] = cyear

            # extract DOI/link from text if exists
            hasLink = cit.find('a')
            if hasLink:
                # get the url and put in citation 
                link = hasLink.get('href')
                citation['link'] = link

            # add to article's list of citations
            citations.append(citation.get('info'))
                
            # insert into citations table
            cur.execute('''INSERT INTO citations (article, info, link, year) VALUES (?, ?, ?, ?)''', 
                        (filename, citation.get('info'), citation.get('link'), citation.get('year')))
            conn.commit()

        # count citations for each article 
        total = len(citations)
        if total == 200:
            total = "200+"      # citations past 200 aren't in html

        # can't put list into table (stupid) so make them strings
        authorString = ', '.join(authors)
        citationString = '\n\n'.join(citations)
        
        # insert into articles table
        cur.execute('''INSERT INTO articles (id, volume, issue, year, title, authors, totalCitations, citations) VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                    (filename, volume, issue, year, title, authorString, total, citationString))
        conn.commit()

cur2 = conn.cursor()
# get number of recent citations for each article
cur.execute('''SELECT id FROM articles''')
for row in cur:
    aId = row[0]
    recentCits = cur2.execute('''SELECT * FROM citations WHERE article = ? AND year >= ?''', (aId, recentYearEnd)).fetchall()
    cur2.execute('''UPDATE articles SET recentCitations = ? WHERE id = ?''', (len(recentCits), aId))
conn.commit() 

print("Collecting volume and issue information...")

# make volumes table
cur.execute('''CREATE TABLE IF NOT EXISTS volumes (
            vol INT,
            year VARCHAR(10),
            totalArticles INT,
            totalCitations INT,
            avgCitPerArticle REAL,
            recentCitations INT
)''')
cur.execute('''DELETE FROM volumes''')
conn.commit()

latest = cur.execute('''SELECT max(volume) FROM articles''').fetchall()
latestVol = latest[0][0]
for i in range(1, latestVol + 1):
    totalCitations = 0
    recentCitations = 0
    totalArticles = 0
    # find all articles in volume
    articlesInVol = cur.execute('''SELECT id, year, recentCitations FROM articles WHERE volume = ? ORDER BY year DESC''', (i,)).fetchall() 
    totalArticles = len(articlesInVol)
    # some volumes are between 2 years because why would it be simple 
    if totalArticles != 0:
        volYear = articlesInVol[0][1]   # set to year from first article to start
        # find all citations for each artice in volume
        for a in articlesInVol:
            citsForVol = cur.execute('''SELECT * FROM citations WHERE article = ?''', (a[0],)).fetchall()
            totalCitations += len(citsForVol)
            # add recent citations
            recentCitations += a[2]
        # calculate avg cits/article    
        avg = round(totalCitations/totalArticles, 3) 
        # put in table 
        cur.execute('''INSERT INTO volumes (vol, year, totalArticles, totalCitations, avgCitPerArticle, recentCitations) 
                    VALUES (?, ?, ?, ?, ?, ?)''', (i, volYear, totalArticles, totalCitations, avg, recentCitations))
conn.commit()

# make issues table
cur.execute('''CREATE TABLE IF NOT EXISTS issues (
            vol INT,
            issue VARCHAR(3),
            year VARCHAR(10),
            totalArticles INT,
            totalCitations INT,
            avgCitPerArticle REAL,
            recentCitations INT
)''')
cur.execute('''DELETE FROM issues''')
conn.commit()

# get each distinct volume-issue combo
distinctIss = cur.execute('''SELECT DISTINCT volume, issue FROM articles''').fetchall()

for i in range(0, len(distinctIss)):
    totalCitations = 0
    recentCitations = 0
    totalArticles = 0
    # find all articles in issue
    articlesInIss = cur.execute('''SELECT id, year, recentCitations FROM articles WHERE volume = ? AND issue = ? ORDER BY year DESC''', (distinctIss[i][0], distinctIss[i][1])).fetchall() 
    totalArticles = len(articlesInIss)
    if totalArticles != 0:
        issYear = articlesInIss[0][1]
        # find all citations for each artice in issue
        for a in articlesInIss:
            citsForIss = cur.execute('''SELECT * FROM citations WHERE article = ?''', (a[0],)).fetchall()
            totalCitations += len(citsForIss)
            # add recent citations
            recentCitations += a[2]
        # calculate avg cits/article    
        avg = round(totalCitations/totalArticles, 3) 
        # put in table 
        cur.execute('''INSERT INTO issues (vol, issue, year, totalArticles, totalCitations, avgCitPerArticle, recentCitations) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)''', (distinctIss[i][0], distinctIss[i][1], issYear, totalArticles, totalCitations, avg, recentCitations))
conn.commit()


outputDir = './output'
if not os.path.exists(outputDir):
    os.mkdir(outputDir)
os.chdir(outputDir)

# get rows and set column names for export
articleRows = cur.execute('''SELECT * FROM articles''').fetchall()
articleColumns = ['id', 'Volume', 'Issue', 'Year', 'Title', 'Authors', 'Total Citations', f'Recent Citations (last {recent} years)', 'Citations']
citationRows = cur.execute('''SELECT * FROM citations''').fetchall()
citationColumns = ['Article id', 'Info', 'Link', 'Year']
volumeRows = cur.execute('''SELECT * FROM volumes''').fetchall()
volumeColumns = ['Volume', 'Year', 'Total articles', 'Total citations', 'Average cit/article', f'Recent citations (last {recent} years)']
issueRows = cur.execute('''SELECT * FROM issues''').fetchall()
issueColumns = ['Volume', 'Issue', 'Year', 'Total articles', 'Total citations', 'Average cit/article', f'Recent citations (last {recent} years)']

# make dataframes
kung = pandas.DataFrame.from_records(articleRows, columns=articleColumns)
fu = pandas.DataFrame.from_records(citationRows, columns=citationColumns)
panda = pandas.DataFrame.from_records(volumeRows, columns=volumeColumns)
two = pandas.DataFrame.from_records(issueRows, columns=issueColumns)

# save to excel file like this so special chars are displayed correctly
# save as 4 sheets in one file
wb = Workbook() 
wsArticles = wb.active 
wsArticles.title = 'Articles'
for r_idx, row in enumerate(dataframe_to_rows(kung, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsArticles.cell(row = r_idx, column = c_idx, value = value)
wsCitations = wb.create_sheet(title = 'Citations')
for r_idx, row in enumerate(dataframe_to_rows(fu, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsCitations.cell(row = r_idx, column = c_idx, value = value)
wsVolumes = wb.create_sheet(title = 'Volumes')
for r_idx, row in enumerate(dataframe_to_rows(panda, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsVolumes.cell(row = r_idx, column = c_idx, value = value)
wsIssues = wb.create_sheet(title = 'Issues')
for r_idx, row in enumerate(dataframe_to_rows(two, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsIssues.cell(row = r_idx, column = c_idx, value = value)
wb.save('CitationRecord.xlsx')
