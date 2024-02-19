import sys 
import os 
import sqlite3
import pandas
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# default number of recent years to analyze and step size
# change this to be input 
recent = 3
step = 5
# need to know current year to know recent years
currentYear = datetime.now().year


conn = sqlite3.connect('jngcitations.db')
cur = conn.cursor()
cur2 = conn.cursor()

# make new tables for charts/graphs 
# need to drop tables if exist so prev data is cleared (might have dif settings)
cur.execute('''DROP TABLE cumulative_total_byvol''')
cur.execute('''CREATE TABLE IF NOT EXISTS cumulative_total_byvol (
            volume INT,
            year INT,
            numArticles INT,
            numCitations INT
)''')
cur.execute('''DELETE FROM cumulative_total_byvol''')
cur.execute('''SELECT vol, year, totalArticles, totalCitations FROM volumes''')
for row in cur:
    vol = row[0]
    yr = row[1]
    articles = row[2]
    citations = row[3]
    cur2.execute('''INSERT INTO cumulative_total_byvol (volume, year, numArticles, numCitations) VALUES (?, ?, ?, ?)''', (vol, yr, articles, citations))
conn.commit()


# calculate total citations for each volume at each time w/ step interval
year = currentYear
while year > 1983:
    colName = 'y' + str(year)
    # add columns to table for each step interval
    cur.execute(f"ALTER TABLE cumulative_total_byvol ADD COLUMN {colName} REAL")
    # get number of volumes
    latest = cur.execute('''SELECT max(volume) FROM articles''').fetchone()
    latestVol = latest[0]
    # loop to calculate for each volume
    for i in range(1, latestVol + 1):
        published = cur.execute('''SELECT year FROM volumes WHERE vol = ?''', (i,)).fetchone()
        if published:
            yrPublished = int(published[0])
            if year >= yrPublished:
                cits = 0
                articles = cur.execute('''SELECT id FROM articles WHERE volume = ?''', (i,)).fetchall()
                for a in articles:
                    articleCits = cur.execute('''SELECT COUNT(*) FROM citations WHERE article = ? AND year <= ?''', (a[0], year)).fetchone()
                    if articleCits:
                        articleCitations = articleCits[0]
                        cits += articleCitations
                cur.execute(f"UPDATE cumulative_total_byvol SET {colName} = ? WHERE volume = ?", (cits, i))
                # print(f"Volume {i}, Year {year}: {cits} citations")
    year -= step
conn.commit()


# make other tables and copy this info there (insert into newtable (select * from cumulative_total_byvol))
# get all columns from cumulative_total_byvol, create table w/ columns, copy info into new table, then calculate avgs
cur.execute('''DROP TABLE cumulative_avg_byvol''')
cur.execute('''CREATE TABLE IF NOT EXISTS cumulative_avg_byvol AS SELECT * FROM cumulative_total_byvol''')
# for each row, update year columns w/ current value divided by numArticles
# get columns
cols = cur.execute('''PRAGMA table_info(cumulative_avg_byvol)''')
columns = [col[1] for col in cols]
for i in range(3, len(columns)):    # start from numCitations (3rd column), calculate avg for each column
    for j in range(1, latestVol):   # loop through each row
        res = cur.execute(f"SELECT numArticles, {columns[i]} FROM cumulative_avg_byvol WHERE volume = {j}").fetchall()
        articles = res[0][0]
        if res[0][1]:    
            cits = res[0][1]
            avg = round(cits/articles, 3) 
            cur.execute(f"UPDATE cumulative_avg_byvol SET {columns[i]} = {avg} WHERE volume = {j}")
conn.commit()


# make article distribution table
cur.execute('''CREATE TABLE IF NOT EXISTS article_distribution (
            u10 INT,
            u25 INT,
            u50 INT,
            u100 INT,
            u200 INT,
            o200 INT
)''')
cur.execute('''DELETE FROM article_distribution''')
under10a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations BETWEEN 0 AND 10''').fetchone()
under25a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations BETWEEN 11 AND 25''').fetchone()
under50a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations BETWEEN 26 AND 50''').fetchone()
under100a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations BETWEEN 51 AND 100''').fetchone()
under200a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations BETWEEN 101 AND 200''').fetchone()
over200a = cur.execute('''SELECT COUNT(*) FROM articles WHERE totalCitations LIKE "200+"''').fetchone()
cur.execute('''INSERT INTO article_distribution (u10, u25, u50, u100, u200, o200) VALUES (?, ?, ?, ?, ?, ?)''', (under10a[0], under25a[0], under50a[0], under100a[0], under200a[0], over200a[0]))
conn.commit()

# make volume distribution table 
cur.execute('''CREATE TABLE IF NOT EXISTS volume_distribution (
            u25 INT,
            u50 INT,
            u100 INT,
            u200 INT,
            u300 INT,
            u400 INT, 
            u500 INT,
            u600 INT,
            u700 INT,
            u800 INT,
            u900 INT,
            u1000 INT,
            o1000 INT
)''')
cur.execute('''DELETE FROM volume_distribution''')
under25v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 0 AND 25''').fetchone()
under50v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 26 AND 50''').fetchone()
under100v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 51 AND 100''').fetchone()
under200v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 101 AND 200''').fetchone()
under300v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 201 AND 300''').fetchone()
under400v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 301 AND 400''').fetchone()
under500v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 401 AND 500''').fetchone()
under600v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 501 AND 600''').fetchone()
under700v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 601 AND 700''').fetchone()
under800v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 701 AND 800''').fetchone()
under900v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 801 AND 900''').fetchone()
under1000v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations BETWEEN 901 AND 1000''').fetchone()
over1000v = cur.execute('''SELECT COUNT(*) FROM volumes WHERE totalCitations > 1000''').fetchone()
cur.execute('''INSERT INTO volume_distribution (u25, u50, u100, u200, u300, u400, u500, u600, u700, u800, u900, u1000, o1000) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (under25v[0], under50v[0], under100v[0], under200v[0], under300v[0], under400v[0], under500v[0], under600v[0], under700v[0], under800v[0], under900v[0], under1000v[0], over1000v[0]))
conn.commit()

# make issue distribution table
cur.execute('''CREATE TABLE IF NOT EXISTS issue_distribution (
            u10 INT,
            u25 INT,
            u50 INT,
            u100 INT,
            u200 INT,
            u300 INT,
            u500 INT,
            o500 INT
)''')
cur.execute('''DELETE FROM issue_distribution''')
under10i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 0 AND 10''').fetchone()
under25i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 11 AND 25''').fetchone()
under50i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 26 AND 50''').fetchone()
under100i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 51 AND 100''').fetchone()
under200i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 101 AND 200''').fetchone()
under300i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 201 AND 300''').fetchone()
under500i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations BETWEEN 301 AND 500''').fetchone()
over500i = cur.execute('''SELECT COUNT(*) FROM issues WHERE totalCitations > 500''').fetchone()
cur.execute('''INSERT INTO issue_distribution(u10, u25, u50, u100, u200, u300, u500, o500) VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', (under10i[0], under25i[0], under50i[0], under100i[0], under200i[0], under300i[0], under500i[0], over500i[0]))
conn.commit()


# get rows and column names for export
totalRows = cur.execute('''SELECT * FROM cumulative_total_byvol''').fetchall()
avgRows = cur.execute('''SELECT * FROM cumulative_avg_byvol''').fetchall()
totalColumns = cur.execute('''PRAGMA table_info(cumulative_total_byvol)''').fetchall()
avgColumns = cur.execute('''PRAGMA table_info(cumulative_avg_byvol)''').fetchall()
totalCols = [col[1] for col in totalColumns]
avgCols = [col[1] for col in avgColumns]

articleRows = cur.execute('''SELECT * FROM article_distribution''').fetchall()
articleColumns = ['0-10', '10-25', '25-50', '50-100', '100-200', '200+']
volumeRows = cur.execute('''SELECT * FROM volume_distribution''').fetchall()
volumeColumns = ['0-25', '25-50', '50-100', '100-200', '200-300', '300-400', '400-500', '500-600', '600-700', '700-800', '800-900', '900-1000', '> 1000']
issueRows = cur.execute('''SELECT * FROM issue_distribution''').fetchall()
issueColumns = ['0-10', '10-25', '25-50', '50-100', '100-200', '200-300', '300-500', '> 500']

# make dataframes
total = pandas.DataFrame.from_records(totalRows, columns=totalCols)
average = pandas.DataFrame.from_records(avgRows, columns=avgCols)
articles = pandas.DataFrame.from_records(articleRows, columns=articleColumns)
volumes = pandas.DataFrame.from_records(volumeRows, columns=volumeColumns)
issues = pandas.DataFrame.from_records(issueRows, columns=issueColumns)

os.chdir('./output')
# save to excel file like this so special chars are displayed correctly
# save as 2 sheets in one file
wb = Workbook() 
wsTotal = wb.active 
wsTotal.title = 'Citations per volume (cumulative)'
for r_idx, row in enumerate(dataframe_to_rows(total, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsTotal.cell(row = r_idx, column = c_idx, value = value)
wsAvg = wb.create_sheet(title = 'Average citations per article (cumulative)')
for r_idx, row in enumerate(dataframe_to_rows(average, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsAvg.cell(row = r_idx, column = c_idx, value = value)
wsArticles = wb.create_sheet(title = 'Citation distribution (articles)')
for r_idx, row in enumerate(dataframe_to_rows(articles, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsArticles.cell(row = r_idx, column = c_idx, value = value)      
wsVolumes = wb.create_sheet(title = 'Citation distribution (volumes)')
for r_idx, row in enumerate(dataframe_to_rows(volumes, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsVolumes.cell(row = r_idx, column = c_idx, value = value)    
wsIssues = wb.create_sheet(title = 'Citation distribution (issues)')
for r_idx, row in enumerate(dataframe_to_rows(issues, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        wsIssues.cell(row = r_idx, column = c_idx, value = value)  
wb.save('CitationRecordOverview.xlsx')
